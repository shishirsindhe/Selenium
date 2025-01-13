from openpyxl import Workbook
import openpyxl
workbook=openpyxl.load_workbook(filename="Demo.xlsx")
sheet=workbook["Demosheet"]
data=[["SHISHIR","SINDHE"],["BILL","GATES"],["STEVE","JOBS"]]
for d in data:
    sheet.append(d)
workbook.save("Demo.xlsx")


wb=openpyxl.load_workbook(filename="Demo.xlsx")
ws=wb['Demosheet']
print(ws["A1"].value)
print(wb["Demosheet"]["B2"].value)
print(ws.cell(row=3,column=1).value)
print(ws.cell(row=5,column=1).value)
